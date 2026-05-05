#!/usr/bin/env python3
"""
AI usage classifier + dual cost recommendations per ai-usage-logic.md (v1.3).

Usage:
  python usage_pipeline.py usage-data.json
  python usage_pipeline.py usage-data.json --out-dir ./out --xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

# --- Step 2: tier rules. First matching substring wins (most specific first; avoids o3-mini matching o3). ---
TIER_RULES_ORDERED: list[tuple[str, float]] = [
    ("gpt-5.5-extra-high", 1.0),
    ("gpt-5.5-high", 0.95),
    ("gpt-5.5-medium", 0.9),
    ("gpt-5.4-xhigh-fast", 0.98),
    ("gpt-5.4-xhigh", 0.97),
    ("gpt-5.4-high-fast", 0.94),
    ("gpt-5.4-high", 0.92),
    ("gpt-5.4-medium", 0.88),
    ("gpt-5.3-codex-xhigh-fast", 0.97),
    ("gpt-5.3-codex-xhigh", 0.95),
    ("gpt-5.3-codex-high-fast", 0.92),
    ("gpt-5.3-codex-high", 0.9),
    ("gpt-5.3-codex-fast", 0.85),
    ("gpt-5.3-codex", 0.88),
    ("gpt-5.2-high", 0.9),
    ("gpt-5.1-codex-max-xhigh", 1.0),
    ("claude-4.6-opus-max-thinking-fast", 1.0),
    ("claude-4.6-opus-max-thinking", 1.0),
    ("claude-4.6-opus-high-thinking-fast", 1.0),
    ("claude-4.6-opus-high-thinking", 1.0),
    ("claude-4.6-opus-max", 1.0),
    ("claude-4.6-opus-high", 1.0),
    ("claude-4.6-sonnet-medium-thinking", 0.78),
    ("claude-4.6-sonnet-medium", 0.76),
    ("claude-4.5-opus-high-thinking", 1.0),
    ("claude-4.5-opus-high", 1.0),
    ("claude-opus-4-7-thinking-max", 1.0),
    ("claude-opus-4-7-thinking-xhigh", 1.0),
    ("claude-opus-4-7-thinking-high", 1.0),
    ("claude-opus-4-7-thinking-medium", 1.0),
    ("claude-opus-4-7-max", 1.0),
    ("claude-opus-4-7-xhigh", 1.0),
    ("claude-opus-4-7-high", 1.0),
    ("claude-opus-4-7-medium", 1.0),
    ("claude-4.5-sonnet-thinking", 0.72),
    ("claude-4.5-sonnet", 0.7),
    ("claude-4-sonnet-thinking", 0.7),
    ("claude-4-sonnet", 0.68),
    ("claude-4.5-haiku-thinking", 0.4),
    ("claude-4.5-haiku", 0.4),
    ("gpt-4o-mini", 0.5),
    ("gpt-4o", 0.85),
    ("gpt-4.1", 0.9),
    ("o3-mini", 0.6),
    ("o3", 1.0),
    ("o1", 1.0),
    ("cursor-slow", 0.85),
    ("cursor-fast", 0.5),
    ("gemini-2.5-pro", 1.0),
    ("gemini-2.5-flash", 0.6),
    ("gemini-2.0-flash", 0.6),
    ("gemini-1.5-flash", 0.35),
    ("gemini-flash", 0.35),
    ("claude-opus", 1.0),
    ("claude-haiku", 0.4),
    ("claude-3.5-haiku", 0.4),
    ("claude-sonnet", 0.65),
    ("haiku", 0.4),
    ("opus", 1.0),
    ("sonnet", 0.65),
]

DEFAULT_TIER = 0.5

MODEL_ALIAS_RULES: list[tuple[str, str]] = [
    ("gpt-5.5-extra-high", "gpt-5.5-extra-high"),
    ("gpt-5.5-high", "gpt-5.5-high"),
    ("gpt-5.5-medium", "gpt-5.5-medium"),
    ("gpt-5.4-xhigh-fast", "gpt-5.4-xhigh-fast"),
    ("gpt-5.4-xhigh", "gpt-5.4-xhigh"),
    ("gpt-5.4-high-fast", "gpt-5.4-high-fast"),
    ("gpt-5.4-high", "gpt-5.4-high"),
    ("gpt-5.4-medium", "gpt-5.4-medium"),
    ("gpt-5.3-codex-xhigh-fast", "gpt-5.3-codex-xhigh-fast"),
    ("gpt-5.3-codex-xhigh", "gpt-5.3-codex-xhigh"),
    ("gpt-5.3-codex-high-fast", "gpt-5.3-codex-high-fast"),
    ("gpt-5.3-codex-high", "gpt-5.3-codex-high"),
    ("gpt-5.3-codex-fast", "gpt-5.3-codex-fast"),
    ("gpt-5.3-codex", "gpt-5.3-codex"),
    ("gpt-5.2-high", "gpt-5.2-high"),
    ("gpt-5.1-codex-max-xhigh", "gpt-5.1-codex-max-xhigh"),
    ("claude-4.6-opus-max-thinking-fast", "claude-4.6-opus-max-thinking-fast"),
    ("claude-4.6-opus-max-thinking", "claude-4.6-opus-max-thinking"),
    ("claude-4.6-opus-high-thinking-fast", "claude-4.6-opus-high-thinking-fast"),
    ("claude-4.6-opus-high-thinking", "claude-4.6-opus-high-thinking"),
    ("claude-4.6-opus-max", "claude-4.6-opus-max"),
    ("claude-4.6-opus-high", "claude-4.6-opus-high"),
    ("claude-4.6-sonnet-medium-thinking", "claude-4.6-sonnet-medium-thinking"),
    ("claude-4.6-sonnet-medium", "claude-4.6-sonnet-medium"),
    ("claude-4.5-opus-high-thinking", "claude-4.5-opus-high-thinking"),
    ("claude-4.5-opus-high", "claude-4.5-opus-high"),
    ("claude-opus-4-7-thinking-max", "claude-opus-4-7-thinking-max"),
    ("claude-opus-4-7-thinking-xhigh", "claude-opus-4-7-thinking-xhigh"),
    ("claude-opus-4-7-thinking-high", "claude-opus-4-7-thinking-high"),
    ("claude-opus-4-7-thinking-medium", "claude-opus-4-7-thinking-medium"),
    ("claude-opus-4-7-max", "claude-opus-4-7-max"),
    ("claude-opus-4-7-xhigh", "claude-opus-4-7-xhigh"),
    ("claude-opus-4-7-high", "claude-opus-4-7-high"),
    ("claude-opus-4-7-medium", "claude-opus-4-7-medium"),
    ("claude-4.5-sonnet-thinking", "claude-4.5-sonnet-thinking"),
    ("claude-4.5-sonnet", "claude-4.5-sonnet"),
    ("claude-4-sonnet-thinking", "claude-4-sonnet-thinking"),
    ("claude-4-sonnet", "claude-4-sonnet"),
    ("claude-4.5-haiku-thinking", "claude-4.5-haiku-thinking"),
    ("claude-4.5-haiku", "claude-4.5-haiku"),
    ("composer-2-fast", "composer-2-fast"),
    ("composer-2", "composer-2"),
    ("gemini-3.1-pro-preview", "gemini-3.1-pro-preview"),
]

# --- Pricing catalog: one candidate per row (ordered: first match wins for model -> tier). ---
@dataclass(frozen=True)
class PriceCandidate:
    key: str
    vendor: str
    label: str
    in_per_m: float
    out_per_m: float
    matchers: tuple[str, ...]


# Matchers ordered most specific first within the same family.
PRICE_CANDIDATES: list[PriceCandidate] = [
    PriceCandidate("gpt-4o-mini", "OpenAI", "GPT-4o mini", 0.15, 0.60, ("gpt-4o-mini",)),
    PriceCandidate("gpt-4o", "OpenAI", "GPT-4o", 2.50, 10.00, ("gpt-4o",)),
    PriceCandidate("gpt-4.1", "OpenAI", "GPT-4.1", 2.00, 8.00, ("gpt-4.1",)),
    PriceCandidate("o3-mini", "OpenAI", "o3-mini", 10.00, 40.00, ("o3-mini",)),  # align with o3 if you prefer
    PriceCandidate("o3", "OpenAI", "o3", 10.00, 40.00, ("o3",)),
    PriceCandidate("o1", "OpenAI", "o1", 15.00, 60.00, ("o1",)),
    PriceCandidate("claude-opus", "Anthropic", "Claude Opus", 15.00, 75.00, ("claude-opus", "claude-opus-")),
    PriceCandidate("opus", "Anthropic", "Claude Opus", 15.00, 75.00, ("opus",)),
    # Haiku before Sonnet so ids like claude-3.5-haiku never match "sonnet"
    PriceCandidate("claude-haiku", "Anthropic", "Claude Haiku", 0.80, 4.00, ("claude-haiku", "claude-3.5-haiku")),
    PriceCandidate("haiku", "Anthropic", "Claude Haiku", 0.80, 4.00, ("haiku",)),
    PriceCandidate("claude-sonnet", "Anthropic", "Claude Sonnet", 3.00, 15.00, ("claude-sonnet",)),
    PriceCandidate("gemini-2.5-pro", "Google", "Gemini 2.5 Pro", 1.25, 10.00, ("gemini-2.5-pro",)),
    PriceCandidate("gemini-2.5-flash", "Google", "Gemini 2.5 Flash", 0.15, 0.60, ("gemini-2.5-flash",)),
    PriceCandidate("gemini-2.0-flash", "Google", "Gemini 2.0 Flash", 0.10, 0.40, ("gemini-2.0-flash",)),
    PriceCandidate("gemini-1.5-flash", "Google", "Gemini 1.5 Flash", 0.10, 0.40, ("gemini-1.5-flash",)),
    PriceCandidate("gemini-flash", "Google", "Gemini Flash", 0.10, 0.40, ("gemini-flash",)),
    PriceCandidate("cursor-slow", "Cursor", "Cursor Slow", 2.50, 10.00, ("cursor-slow",)),
    PriceCandidate("cursor-fast", "Cursor", "Cursor Fast", 0.15, 0.60, ("cursor-fast",)),
]

UNKNOWN_PRICE = (2.50, 10.00)


def norm_model(model: str) -> str:
    return model.strip().lower()


def parse_cursor_timestamp(ts: Any) -> str:
    """Normalize Cursor event timestamps to YYYY-MM-DD."""
    if ts is None:
        return ""
    s = str(ts).strip()
    if not s:
        return ""
    # Cursor may return ms epoch as string or ISO 8601 in different endpoints.
    try:
        if s.isdigit():
            n = int(s)
            if n > 10_000_000_000:  # milliseconds
                return datetime.fromtimestamp(n / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
            return datetime.fromtimestamp(n, tz=timezone.utc).strftime("%Y-%m-%d")
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        return s[:10]


def cursor_cache_tokens(row: dict[str, Any]) -> int:
    tok = row.get("tokenUsage") if isinstance(row.get("tokenUsage"), dict) else {}
    return int(tok.get("cacheReadTokens") or 0) + int(tok.get("cacheWriteTokens") or 0)


def cursor_cost_usd(row: dict[str, Any]) -> float:
    tok = row.get("tokenUsage") if isinstance(row.get("tokenUsage"), dict) else {}
    charged_cents = row.get("chargedCents")
    total_cents = tok.get("totalCents")
    cents = charged_cents if charged_cents is not None else total_cents
    try:
        return float(cents or 0.0) / 100.0
    except Exception:
        return 0.0


def cursor_request_row_to_usage_row(row: dict[str, Any]) -> dict[str, Any]:
    tok = row.get("tokenUsage") if isinstance(row.get("tokenUsage"), dict) else {}
    is_token_based = bool(row.get("isTokenBasedCall"))
    cached = cursor_cache_tokens(row)
    tool_turns = 1 if row.get("isHeadless") else 0
    requests_costs = row.get("requestsCosts")
    try:
        requests_costs_val = float(requests_costs or 0.0)
    except Exception:
        requests_costs_val = 0.0
    return {
        "date": parse_cursor_timestamp(row.get("timestamp")),
        "provider": "Cursor",
        "model": row.get("model") or "",
        "user_email": row.get("userEmail") or "",
        "requests_count": 1,
        "input_tokens": int(tok.get("inputTokens") or 0),
        "output_tokens": int(tok.get("outputTokens") or 0),
        "cached_tokens": cached,
        "cost_usd": cursor_cost_usd(row),
        "tool_turns": tool_turns,
        "web_search_requests": 0,
        "cursor_kind": row.get("kind") or "",
        "cursor_max_mode": bool(row.get("maxMode")),
        "cursor_chargeable": bool(row.get("isChargeable")),
        "cursor_headless": bool(row.get("isHeadless")),
        "cursor_is_token_based": is_token_based,
        "cursor_requests_costs": requests_costs_val,
        "cursor_token_fee": float(row.get("cursorTokenFee") or 0.0),
    }


def is_cursor_filtered_usage_events(raw: Any) -> bool:
    return isinstance(raw, dict) and "usageEvents" in raw and isinstance(raw.get("usageEvents"), list)


def model_tier(model: str) -> float:
    m = norm_model(model)
    for sub, t in TIER_RULES_ORDERED:
        if sub in m:
            return t
    return DEFAULT_TIER


def resolve_price_candidate(model: str) -> PriceCandidate | None:
    m = norm_model(model)
    for alias, canonical in MODEL_ALIAS_RULES:
        if alias in m:
            for c in PRICE_CANDIDATES:
                if c.key == canonical:
                    return c
    for c in PRICE_CANDIDATES:
        for pat in c.matchers:
            if pat in m:
                return c
    return None


def blended_cost_per_1m(
    tin: int, tout: int, in_per_m: float, out_per_m: float
) -> float:
    tot = tin + tout
    if tot <= 0:
        return 0.0
    return (in_per_m * tin + out_per_m * tout) / tot


def normalize_provider(p: str | None) -> str | None:
    if not p:
        return None
    s = p.strip().lower()
    mapping = {
        "openai": "OpenAI",
        "anthropic": "Anthropic",
        "google": "Google",
        "google ai": "Google",
        "cursor": "Cursor",
    }
    return mapping.get(s, p.strip().title())


def vendor_from_model(model: str) -> str | None:
    """First-match vendor map (spec Step 4)."""
    m = norm_model(model)
    rules: list[tuple[str, str]] = [
        ("gpt-5.5", "OpenAI"),
        ("gpt-5.4", "OpenAI"),
        ("gpt-5.3", "OpenAI"),
        ("gpt-5.2", "OpenAI"),
        ("gpt-5.1", "OpenAI"),
        ("gpt-", "OpenAI"),
        ("o1", "OpenAI"),
        ("o3", "OpenAI"),
        ("claude-4.6", "Anthropic"),
        ("claude-4.5", "Anthropic"),
        ("claude-4-sonnet", "Anthropic"),
        ("claude-opus-4-7", "Anthropic"),
        ("claude-", "Anthropic"),
        ("gemini-3.1", "Google"),
        ("gemini-2.5", "Google"),
        ("gemini-2.0", "Google"),
        ("gemini-1.5", "Google"),
        ("gemini", "Google"),
        ("composer-", "Cursor"),
        ("cursor-", "Cursor"),
        ("default", "Cursor"),
    ]
    for sub, v in rules:
        if sub in m:
            return v
    return None


def row_vendor(provider: str | None, model: str) -> str | None:
    v = normalize_provider(provider)
    if v in ("OpenAI", "Anthropic", "Google", "Cursor"):
        return v
    return vendor_from_model(model)


def classify_category(
    *,
    reqs: int,
    rpd: float,
    avg_in: float,
    ratio: float,
    tier: float,
    model: str,
    tools: int,
    wsr: int,
    provider_norm: str | None,
) -> str:
    m = norm_model(model)
    if reqs < 20 and rpd < 1.5:
        return "🧪 Explorer"
    if "cursor" in m or provider_norm == "Cursor":
        return "🧑‍💻 Power / Technical"
    if "opus" in m:
        return "🧑‍💻 Power / Technical"
    if tools >= 8 or wsr >= 6:
        return "🧑‍💻 Power / Technical"
    if (
        rpd > 3.6
        or avg_in > 15_000
        or ("sonnet" in m and (avg_in > 10_000 or rpd > 3.2))
        or (tier >= 0.85 and reqs > 50)
    ):
        return "🧑‍💻 Power / Technical"
    if avg_in < 6_500 and ratio < 0.73:
        return "🔍 Lookup / Q&A"
    if ratio > 0.76 and avg_in < 12_000:
        return "✍️ Content Generator"
    if 1.5 <= rpd <= 3.6 and 6_500 <= avg_in <= 15_000:
        return "💬 Conversational"
    return "💬 Conversational"


def _savings_percent_of_row_spend(save_usd: float, total_cost_usd: float) -> float | None:
    if total_cost_usd and total_cost_usd > 0:
        return 100.0 * save_usd / total_cost_usd
    return None


def _fmt_money(x: float) -> str:
    return f"${x:,.2f}"


def _fmt_savings_phrase(save_usd: float, pct: float | None) -> str:
    """Human-readable savings line (percent + dollars)."""
    money = _fmt_money(save_usd)
    if pct is None:
        return f"roughly {money} less on this user-and-model total at list prices"
    return f"roughly {pct:.0f}% lower modeled cost ({money} less on this user-and-model total at list prices)"


def plain_english_why_cheaper(
    *,
    category: str,
    raw_model: str,
    current_friendly: str,
    total_cost_usd: float,
    ok_global: bool,
    global_cand: PriceCandidate | None,
    save_global: float,
    ok_same: bool,
    same_cand: PriceCandidate | None,
    save_same: float,
    vendor_name: str | None,
) -> str:
    """Short rationale for non-technical readers (also written to JSON/CSV)."""
    if "Explorer" in category:
        angle = (
            "Your usage is still fairly light, so you often do not need the most expensive model tier "
            "to get useful answers."
        )
    elif "Power / Technical" in category:
        angle = (
            "You are a heavy user with large or frequent requests, so small differences in per-token "
            "price add up to large amounts over a month."
        )
    elif "Lookup / Q&A" in category:
        angle = (
            "Your pattern looks like short questions and quick answers—similar to search or help-desk use—"
            "where top-tier models are usually hard to justify on cost."
        )
    elif "Content Generator" in category:
        angle = (
            "You produce a lot of model output, so output-token pricing dominates your bill; "
            "a cheaper model that still passes your quality bar can cut spend sharply."
        )
    elif "Conversational" in category:
        angle = (
            "Your usage is steady back-and-forth conversation; many teams find mid-tier or faster "
            "economy models enough after a short quality check."
        )
    else:
        angle = "Your usage may not require the highest-priced model on every request."

    parts: list[str] = [angle]

    parts.append(f'You are on {current_friendly} (model id "{raw_model}").')


    parts.append(
        "The following options use public list prices and your recorded input/output mix; "
        "each is at least roughly 5% cheaper than your current tier for that mix."
    )

    if ok_global and global_cand:
        pg = _savings_percent_of_row_spend(save_global, total_cost_usd)
        parts.append(
            f"If you can switch providers, {global_cand.vendor} {global_cand.label} is the cheapest option we catalog: "
            f"at list prices that would be {_fmt_savings_phrase(save_global, pg)} "
            "assuming your prompts and quality needs stay the same—validate with a pilot before committing."
        )

    if ok_same and same_cand and vendor_name:
        ps = _savings_percent_of_row_spend(save_same, total_cost_usd)
        if ok_global and global_cand and global_cand.key == same_cand.key:
            parts.append(
                "The best choice for your current vendor matches the global cheapest option, "
                "so one move satisfies both staying within your vendor rules and minimizing cost."
            )
        else:
            parts.append(
                f"If you must stay with {vendor_name}, {same_cand.label} is the cheapest qualifying model "
                f"we still list for that vendor: {_fmt_savings_phrase(save_same, ps)} "
                "compared with your current model, subject to the same quality checks."
            )

    return " ".join(parts)


def pick_recommendation(
    *,
    tin: int,
    tout: int,
    total_cost: float,
    current_cand: PriceCandidate | None,
    current_in_m: float,
    current_out_m: float,
    vendor_filter: str | None,
) -> tuple[PriceCandidate | None, bool, float]:
    """
    Return (best_candidate, is_cheaper, est_savings_usd).
    vendor_filter None = global; else same-vendor only.
    """
    tot = tin + tout
    if tot <= 0:
        return None, False, 0.0

    curr_per_m = blended_cost_per_1m(tin, tout, current_in_m, current_out_m)
    if curr_per_m <= 0:
        return None, False, 0.0

    best: PriceCandidate | None = None
    best_per_m: float = float("inf")
    best_label = ""

    for c in PRICE_CANDIDATES:
        if vendor_filter is not None and c.vendor != vendor_filter:
            continue
        if current_cand is not None and c.key == current_cand.key:
            continue
        per_m = blended_cost_per_1m(tin, tout, c.in_per_m, c.out_per_m)
        if per_m < curr_per_m * 0.95:
            label = f"{c.vendor} — {c.label}"
            if per_m < best_per_m or (
                math.isclose(per_m, best_per_m) and label < best_label
            ):
                best = c
                best_per_m = per_m
                best_label = label

    if best is None:
        return None, False, 0.0

    savings_pct = (1.0 - best_per_m / curr_per_m) * 100.0
    est_savings = total_cost * (savings_pct / 100.0)
    return best, True, est_savings


def unwrap_records(raw: Any) -> list[dict[str, Any]]:
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        for k in ("data", "rows", "records", "results"):
            if k in raw and isinstance(raw[k], list):
                return raw[k]
    raise ValueError("Expected JSON array or object with data/rows/records/results list")


def iter_normalized_records(rows: list[dict[str, Any]]) -> Iterator[dict[str, Any]]:
    for r in rows:
        d = {str(k).lower().replace(" ", "_"): v for k, v in r.items()}
        yield d


@dataclass
class Agg:
    user_email: str
    model: str
    user_name: str | None = None
    department: str | None = None
    team: str | None = None
    provider: str | None = None
    total_requests: int = 0
    total_input: int = 0
    total_output: int = 0
    total_cached: int = 0
    total_cost: float = 0.0
    tools: int = 0
    wsr: int = 0
    cursor_kind_counts: dict[str, int] = field(default_factory=lambda: defaultdict(int))
    cursor_headless_count: int = 0
    cursor_chargeable_count: int = 0
    cursor_token_based_count: int = 0
    cursor_requests_costs: float = 0.0
    cursor_token_fee: float = 0.0
    dates: set[str] = field(default_factory=set)


def aggregate(rows: list[dict[str, Any]]) -> list[Agg]:
    groups: dict[tuple[str, str], Agg] = {}
    for d in iter_normalized_records(rows):
        email = d.get("user_email")
        model = d.get("model")
        if not email or not model:
            continue
        key = (str(email).strip(), str(model).strip())
        g = groups.get(key)
        if g is None:
            g = Agg(user_email=key[0], model=key[1])
            groups[key] = g
        rc = int(d.get("requests_count") or 1)
        g.total_requests += rc
        g.total_input += int(d.get("input_tokens") or 0)
        g.total_output += int(d.get("output_tokens") or 0)
        g.total_cached += int(d.get("cached_tokens") or 0)
        g.total_cost += float(d.get("cost_usd") or 0.0)
        g.tools += int(d.get("tool_turns") or 0)
        g.wsr += int(d.get("web_search_requests") or 0)
        kind = d.get("cursor_kind")
        if kind:
            g.cursor_kind_counts[str(kind)] += 1
        if d.get("cursor_headless"):
            g.cursor_headless_count += 1
        if d.get("cursor_chargeable"):
            g.cursor_chargeable_count += 1
        if d.get("cursor_is_token_based"):
            g.cursor_token_based_count += 1
        g.cursor_requests_costs += float(d.get("cursor_requests_costs") or 0.0)
        g.cursor_token_fee += float(d.get("cursor_token_fee") or 0.0)
        dt = d.get("date")
        if dt:
            ds = str(dt)
            g.dates.add(ds[:10] if len(ds) >= 10 else ds)
        prov = normalize_provider(d.get("provider"))
        if prov and g.provider is None:
            g.provider = prov
        elif prov and g.provider and prov != g.provider:
            pass  # keep first
        un = d.get("user_name")
        if un and not g.user_name:
            g.user_name = str(un)
        dep = d.get("department")
        if dep and not g.department:
            g.department = str(dep)
        tm = d.get("team")
        if tm and not g.team:
            g.team = str(tm)
    return list(groups.values())


def process_agg(g: Agg) -> dict[str, Any]:
    tr = g.total_requests
    tin, tout = g.total_input, g.total_output
    active_days = len(g.dates) if g.dates else max(1, tr)
    rpd = tr / active_days if active_days else 0.0
    avg_in = tin / tr if tr else 0.0
    avg_out = tout / tr if tr else 0.0
    ratio = (tout / tin) if tin > 0 else 1.0
    tier = model_tier(g.model)
    pv = row_vendor(g.provider, g.model)

    cat = classify_category(
        reqs=tr,
        rpd=rpd,
        avg_in=avg_in,
        ratio=ratio,
        tier=tier,
        model=g.model,
        tools=g.tools,
        wsr=g.wsr,
        provider_norm=pv,
    )

    cur_c = resolve_price_candidate(g.model)
    if cur_c:
        c_in, c_out = cur_c.in_per_m, cur_c.out_per_m
    else:
        c_in, c_out = UNKNOWN_PRICE

    g_cand, ok_g, save_g = pick_recommendation(
        tin=tin,
        tout=tout,
        total_cost=g.total_cost,
        current_cand=cur_c,
        current_in_m=c_in,
        current_out_m=c_out,
        vendor_filter=None,
    )
    sv = row_vendor(g.provider, g.model)
    b_cand, ok_b, save_b = pick_recommendation(
        tin=tin,
        tout=tout,
        total_cost=g.total_cost,
        current_cand=cur_c,
        current_in_m=c_in,
        current_out_m=c_out,
        vendor_filter=sv,
    )

    def fmt_cand(c: PriceCandidate | None) -> str:
        if c is None:
            return "—"
        return f"{c.vendor} — ➡️ {c.label}"

    expl_parts = [f"Category: {cat}."]
    if ok_g:
        expl_parts.append(f"Global best: {fmt_cand(g_cand)}.")
    else:
        expl_parts.append("No ≥5% cheaper model in full catalog.")
    if sv:
        if ok_b:
            expl_parts.append(f"Same vendor ({sv}): {fmt_cand(b_cand)}.")
        else:
            expl_parts.append(f"Same vendor ({sv}): already on cheapest listed tier or no ≥5% cheaper option.")
    else:
        expl_parts.append("Same-vendor filter N/A (vendor unknown).")

    current_friendly = cur_c.label if cur_c else g.model
    why_plain = plain_english_why_cheaper(
        category=cat,
        raw_model=g.model,
        current_friendly=current_friendly,
        total_cost_usd=g.total_cost,
        ok_global=ok_g,
        global_cand=g_cand,
        save_global=save_g,
        ok_same=ok_b if sv else False,
        same_cand=b_cand,
        save_same=save_b,
        vendor_name=sv,
    )

    return {
        "user_email": g.user_email,
        "user_name": g.user_name,
        "department": g.department,
        "team": g.team,
        "provider": g.provider,
        "model": g.model,
        "total_requests": tr,
        "total_input": tin,
        "total_output": tout,
        "total_cached": g.total_cached,
        "total_cost_usd": round(g.total_cost, 6),
        "active_days": active_days,
        "requests_per_day": round(rpd, 4),
        "avg_input": round(avg_in, 2),
        "avg_output": round(avg_out, 2),
        "ratio": round(ratio, 4),
        "model_tier": tier,
        "cache_rate": round(g.total_cached / tin, 4) if tin else 0.0,
        "category": cat,
        "recommendation_global": fmt_cand(g_cand),
        "is_cheaper_global": ok_g,
        "est_savings_global_usd": round(save_g, 4) if ok_g else None,
        "recommendation_same_vendor": fmt_cand(b_cand) if sv else "—",
        "is_cheaper_same_vendor": ok_b if sv else False,
        "est_savings_same_vendor_usd": round(save_b, 4) if ok_b and sv else None,
        "cursor_kind_counts": dict(sorted(g.cursor_kind_counts.items())),
        "cursor_headless_count": g.cursor_headless_count,
        "cursor_chargeable_count": g.cursor_chargeable_count,
        "cursor_token_based_count": g.cursor_token_based_count,
        "cursor_requests_costs": round(g.cursor_requests_costs, 6),
        "cursor_token_fee": round(g.cursor_token_fee, 6),
        "explanation": " ".join(expl_parts),
        "why_cheaper_plain_english": why_plain,
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("")
        return
    keys = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(
    path: Path,
    all_rows: list[dict[str, Any]],
    opportunities: list[dict[str, Any]],
    summary: list[dict[str, Any]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    wb = Workbook()
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ws0 = wb.active
    ws0.title = "Executive"
    total_cost = sum(r.get("total_cost_usd") or 0 for r in all_rows)
    sg = sum(
        (r.get("est_savings_global_usd") or 0)
        for r in all_rows
        if r.get("is_cheaper_global")
    )
    sb = sum(
        (r.get("est_savings_same_vendor_usd") or 0)
        for r in all_rows
        if r.get("is_cheaper_same_vendor")
    )
    ws0.append(["Metric", "Value"])
    ws0.append(["Total cost (opportunity rows only)", total_cost])
    ws0.append(["Est. savings (any vendor)", sg])
    ws0.append(["Est. savings (same vendor)", sb])
    if all_rows and "cursor_requests_costs" in all_rows[0]:
        ws0.append(["Cursor requests costs", sum(r.get("cursor_requests_costs") or 0 for r in all_rows)])
        ws0.append(["Cursor token fee", sum(r.get("cursor_token_fee") or 0 for r in all_rows)])

    ws1 = wb.create_sheet("Opportunities")
    if opportunities:
        headers = list(opportunities[0].keys())
        ws1.append(headers)
        for r in opportunities:
            ws1.append([r.get(h) for h in headers])
        # green last savings columns
        ig = headers.index("est_savings_global_usd") + 1 if "est_savings_global_usd" in headers else None
        ib = headers.index("est_savings_same_vendor_usd") + 1 if "est_savings_same_vendor_usd" in headers else None
        for ri, r in enumerate(opportunities, start=2):
            if ig and r.get("est_savings_global_usd") is not None:
                ws1.cell(row=ri, column=ig).fill = green
            if ib and r.get("est_savings_same_vendor_usd") is not None:
                ws1.cell(row=ri, column=ib).fill = green

    ws2 = wb.create_sheet("Summary")
    if summary:
        ws2.append(list(summary[0].keys()))
        for r in summary:
            ws2.append([r.get(k) for k in summary[0].keys()])

    ws3 = wb.create_sheet("Legend")
    ws3.append(["Category", "Description"])
    for line in [
        ("🧑‍💻 Power / Technical", "Heavy usage, large inputs, flagship or IDE models, tools/search."),
        ("✍️ Content Generator", "High output vs input — drafting, writing."),
        ("💬 Conversational", "Balanced back-and-forth."),
        ("🔍 Lookup / Q&A", "Short prompts and outputs, many requests."),
        ("🧪 Explorer", "Low or irregular usage."),
    ]:
        ws3.append(list(line))

    wb.save(path)


def summary_by_category(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by[r["category"]].append(r)
    out = []
    for cat, lst in sorted(by.items()):
        users = len({r["user_email"] for r in lst})
        cost = sum(r["total_cost_usd"] for r in lst)
        # mode of global recommendation string among those with savings
        modes_g = [r["recommendation_global"] for r in lst if r.get("is_cheaper_global")]
        modes_b = [r["recommendation_same_vendor"] for r in lst if r.get("is_cheaper_same_vendor")]
        out.append(
            {
                "category": cat,
                "users": users,
                "total_cost_usd": round(cost, 2),
                "top_global_target_mode": max(set(modes_g), key=modes_g.count) if modes_g else "—",
                "top_same_vendor_target_mode": max(set(modes_b), key=modes_b.count) if modes_b else "—",
            }
        )
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Classify usage + dual cost recommendations.")
    ap.add_argument("input_json", type=Path, help="Path to usage JSON (array or wrapped)")
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=Path("out"),
        help="Output directory (default: ./out)",
    )
    ap.add_argument("--xlsx", action="store_true", help="Write usage-report.xlsx (needs openpyxl)")
    args = ap.parse_args()

    raw = json.loads(args.input_json.read_text(encoding="utf-8"))
    if is_cursor_filtered_usage_events(raw):
        records = [cursor_request_row_to_usage_row(r) for r in raw.get("usageEvents", [])]
    else:
        records = unwrap_records(raw)
    aggs = aggregate(records)
    processed = [process_agg(g) for g in aggs]
    processed.sort(key=lambda r: (-(r.get("total_cost_usd") or 0), r["user_email"], r["model"]))

    opportunities = [
        r
        for r in processed
        if r.get("is_cheaper_global") or r.get("is_cheaper_same_vendor")
    ]
    opportunities.sort(
        key=lambda r: (
            -(r.get("est_savings_global_usd") or 0),
            -(r.get("est_savings_same_vendor_usd") or 0),
        )
    )

    args.out_dir.mkdir(parents=True, exist_ok=True)
    all_path = args.out_dir / "usage-report-all.json"
    opp_path = args.out_dir / "usage-report-opportunities.json"
    all_path.write_text(json.dumps(processed, indent=2), encoding="utf-8")
    opp_path.write_text(json.dumps(opportunities, indent=2), encoding="utf-8")
    write_csv(args.out_dir / "usage-report-all.csv", processed)
    write_csv(args.out_dir / "usage-report-opportunities.csv", opportunities)

    summ = summary_by_category(processed)
    (args.out_dir / "usage-summary-by-category.json").write_text(
        json.dumps(summ, indent=2), encoding="utf-8"
    )

    if args.xlsx:
        write_xlsx(args.out_dir / "usage-report.xlsx", processed, opportunities, summ)

    print(f"Loaded {len(records)} raw rows → {len(processed)} user×model aggregates.")
    print(f"Opportunity rows (either track saves ≥5%): {len(opportunities)}")
    print(f"Wrote: {all_path}, {opp_path}, CSVs under {args.out_dir}")
    if args.xlsx:
        print(f"Wrote: {args.out_dir / 'usage-report.xlsx'}")


if __name__ == "__main__":
    main()
